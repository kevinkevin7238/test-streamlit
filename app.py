import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import io

st.title("Time Study CSV to Excel Formatter")

# Upload CSV
csv_file = st.file_uploader("Upload your time_study_data.csv", type="csv")
# Upload Excel template
excel_template = st.file_uploader("Upload your Book.xlsx template", type="xlsx")

if csv_file and excel_template:
    try:
        raw_data = pd.read_csv(csv_file)

        # Convert Element in the order it appears
        raw_data['Element'] = pd.Categorical(raw_data['Element'], 
                                             ordered=True, 
                                             categories=raw_data['Element'].unique())
        
        # Group Element + Time
        grouped = raw_data.groupby('Element', observed=True)['Time (IM)'].apply(list).reset_index()

        # Format the data
        formatted_rows = []
        for _, row in grouped.iterrows():
            element_times = row['Time (IM)']
            element = row['Element']
            first_row = [element] + [None]*5
            if len(element_times) == 1:
                first_row += [element_times[0]] + [None]*10
                second_row = [None]*16
            else:
                first_row += [None]
                times_first_row = element_times[:10]
                times_second_row = element_times[10:20]
                first_row += times_first_row + [None]*(10-len(times_first_row))
                second_row = [None]*6 + times_second_row + [None]*(10-len(times_second_row))
            
            formatted_rows.append(first_row)
            formatted_rows.append(second_row if len(element_times) > 10 else [None]*16)

        column_names = ['ELEMENT DESCRIPTION'] + ['']*5 + ['ALL'] + [str(i) for i in range(1,11)]
        formatted_data = pd.DataFrame(formatted_rows, columns=column_names)

        # Load template
        wb = load_workbook(excel_template)
        ws = wb.active

        # Start from B20
        for r_idx, row in enumerate(dataframe_to_rows(formatted_data, index=False, header=False), 20):
            for c_idx, value in enumerate(row, 2):
                cell = ws.cell(row=r_idx, column=c_idx)
                # Only write if the cell is not merged or it's the top-left of a merged range
                if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    cell.value = value


        # Merge cells for element description
        for i in range(0, len(formatted_rows), 2):
            start_row = 20 + i
            ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=6)
            ws.cell(row=start_row, column=2).value = formatted_rows[i][0]

        # Save to a BytesIO object for download
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        st.success("Excel file processed successfully!")
        st.download_button(
            label="Download Processed Excel",
            data=output,
            file_name="Processed_Book.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"An error occurred: {e}")
